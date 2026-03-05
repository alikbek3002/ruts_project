from __future__ import annotations

from datetime import datetime, date, timedelta

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.core.deps import get_current_user, require_role
from app.core.monitor import timed
from app.core.provisioning import password_fingerprint
from app.core.security import hash_password
from app.db.supabase_client import get_supabase

router = APIRouter()
ABSENCE_ATTENDANCE_TYPES = {"absent", "excused", "sick"}


def _timetable_entries_for_teacher(sb, teacher_id: str, select_fields: str, weekday: int | None = None) -> list[dict]:
    q1 = sb.table("timetable_entries").select(select_fields).eq("active", True).eq("teacher_id", teacher_id)
    if weekday is not None:
        q1 = q1.eq("weekday", weekday)
    return q1.execute().data or []


def _timetable_entries_for_class(
    sb,
    class_id: str,
    select_fields: str,
    teacher_id: str | None = None,
) -> list[dict]:
    """
    Backward-compatible class timetable fetch:
    - new schema: class_ids contains class_id
    - legacy schema/rows: class_id equals class_id
    """
    q_multi = (
        sb.table("timetable_entries")
        .select(select_fields)
        .eq("active", True)
        .cs("class_ids", [class_id])
    )
    if teacher_id:
        q_multi = q_multi.eq("teacher_id", teacher_id)
    rows_multi = q_multi.execute().data or []

    q_legacy = (
        sb.table("timetable_entries")
        .select(select_fields)
        .eq("active", True)
        .eq("class_id", class_id)
    )
    if teacher_id:
        q_legacy = q_legacy.eq("teacher_id", teacher_id)
    rows_legacy = q_legacy.execute().data or []

    by_id: dict[str, dict] = {}
    for row in rows_multi + rows_legacy:
        rid = row.get("id")
        if not rid:
            continue
        by_id[str(rid)] = row
    return list(by_id.values())


def _normalize_lesson_date(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if "T" in s:
        s = s.split("T", 1)[0]
    return s[:10] if len(s) >= 10 else s


def _entry_matches_weekday(entry_weekday: object, target_weekday: int) -> bool:
    """
    Compatibility for both weekday conventions:
    - 0..6 (Mon..Sun) [current]
    - 1..7 (Mon..Sun) [legacy]
    """
    if entry_weekday is None:
        return False
    try:
        w = int(str(entry_weekday))
    except (ValueError, TypeError):
        return False
    if 0 <= w <= 6:
        return w == target_weekday
    if 1 <= w <= 7:
        return (w - 1) == target_weekday
    return False


def _ensure_legacy_student_id_for_enrollment(sb, class_id: str, enrollment: dict) -> str | None:
    """
    Ensure class enrollment has legacy_student_id.
    If missing, create a technical inactive student user and bind it.
    """
    legacy = enrollment.get("legacy_student_id")
    if legacy:
        return str(legacy)

    enrollment_id = enrollment.get("id")
    if not enrollment_id:
        return None
    enrollment_id = str(enrollment_id)

    username = f"enr-{enrollment_id}"
    full_name = (enrollment.get("student_full_name") or "").strip()
    if not full_name:
        num = enrollment.get("student_number")
        full_name = f"Student #{num}" if num is not None else f"Student {enrollment_id[:8]}"

    existing = (
        sb.table("users")
        .select("id")
        .eq("username", username)
        .limit(1)
        .execute()
        .data
        or []
    )
    if existing:
        user_id = str(existing[0].get("id"))
    else:
        raw_password = f"enrollment:{enrollment_id}"
        payload = {
            "role": "student",
            "username": username,
            "full_name": full_name,
            "password_hash": hash_password(raw_password),
            "password_fingerprint": password_fingerprint(raw_password),
            "must_change_password": False,
            "is_active": False,
        }
        try:
            created = sb.table("users").insert(payload).execute().data or []
        except Exception:
            # Backward-compatible fallback if password_fingerprint column is absent.
            payload.pop("password_fingerprint", None)
            created = sb.table("users").insert(payload).execute().data or []
        if not created:
            raise HTTPException(status_code=500, detail="Failed to provision student account")
        user_id = str(created[0].get("id"))

    sb.table("class_enrollments").update({"legacy_student_id": user_id}).eq("class_id", class_id).eq("id", enrollment_id).execute()
    return user_id


def _resolve_legacy_student_id_for_class(sb, class_id: str, raw_student_id: str) -> str | None:
    sid = str(raw_student_id)

    by_legacy = (
        sb.table("class_enrollments")
        .select("id,legacy_student_id,student_full_name,student_number")
        .eq("class_id", class_id)
        .eq("legacy_student_id", sid)
        .limit(1)
        .execute()
        .data
        or []
    )
    if by_legacy and by_legacy[0].get("legacy_student_id"):
        return str(by_legacy[0]["legacy_student_id"])

    by_enrollment = (
        sb.table("class_enrollments")
        .select("id,legacy_student_id,student_full_name,student_number")
        .eq("class_id", class_id)
        .eq("id", sid)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not by_enrollment:
        return None

    enrollment = by_enrollment[0]
    if enrollment.get("legacy_student_id"):
        return str(enrollment["legacy_student_id"])

    return _ensure_legacy_student_id_for_enrollment(sb, class_id, enrollment)


def _teacher_can_access_entry(sb, teacher_id: str, entry_data: dict) -> bool:
    return entry_data.get("teacher_id") == teacher_id


def _lesson_journal_makeup_supported(sb) -> bool:
    try:
        sb.table("lesson_journal").select("makeup_grade,attendance_makeup").limit(1).execute()
        return True
    except Exception:
        return False


class GradeEntry(BaseModel):
    subject: str
    grade: int
    comment: str | None = None


class SaveGradesIn(BaseModel):
    student_id: str
    grades: list[GradeEntry]


class AddGradeIn(BaseModel):
    student_id: str
    timetable_entry_id: str
    lesson_date: str  # YYYY-MM-DD
    grade: int | None = None  # None = no grade, 2-5 valid values
    makeup_grade: int | None = None  # Additional grade after rework (keeps original grade)
    present: bool | None = None
    comment: str | None = None
    attendance_type: str | None = None  # present, absent, duty (Кезмет), excused (Арыз), sick (Оруу)
    attendance_makeup: bool | None = None  # True = absence was worked off (ОТР)


@router.get("/teacher/classes")
@timed("get_teacher_classes")
def get_teacher_classes(user: dict = require_role("teacher", "admin", "manager")):
    """Получить все классы (для выбора журнала)"""
    sb = get_supabase()
    
    # Return all classes, sorted by name
    # We might want to optimize this if there are thousands, but for now it's fine.
    # We mark "my" classes (where teacher has schedule) for UI highlighting if needed.
    
    # 1. Get all classes
    all_classes = sb.table("classes").select("id,name").order("name").execute().data or []
    
    # 2. Get my classes (to highlight or sort to top?)
    # For now, just return all. The user asked for "just all classes".
    
    return {"classes": all_classes}


@router.get("/classes/{class_id}/subjects")
def get_class_subjects(class_id: str, user: dict = require_role("teacher", "admin", "manager")):
    """Return class subjects from active timetable entries."""
    sb = get_supabase()
    rows = _timetable_entries_for_class(
        sb,
        class_id,
        "id,subject_id,subject,teacher_id",
        user["id"] if user.get("role") == "teacher" else None,
    )

    subject_ids = sorted({str(r.get("subject_id")) for r in rows if r.get("subject_id")})
    names_by_id: dict[str, str] = {}
    if subject_ids:
        subject_rows = (
            sb.table("subjects")
            .select("id,name")
            .in_("id", subject_ids)
            .is_("archived_at", "null")
            .execute()
            .data
            or []
        )
        names_by_id = {
            str(s.get("id")): str(s.get("name") or "").strip()
            for s in subject_rows
            if s.get("id")
        }

    subjects_map: dict[str, dict] = {}
    for r in rows:
        sid_raw = r.get("subject_id")
        if not sid_raw:
            continue
        sid = str(sid_raw)
        if sid in subjects_map:
            continue
        name = names_by_id.get(sid) or str(r.get("subject") or "").strip()
        if not name:
            continue
        subjects_map[sid] = {
            "id": sid,
            "name": name,
            "is_mine": bool(r.get("teacher_id") == user.get("id")),
        }

    sorted_subjects = sorted(
        subjects_map.values(),
        key=lambda x: (not x.get("is_mine", False), x.get("name") or ""),
    )

    return {"subjects": sorted_subjects}

@router.get("/teacher/schedule")
@timed("get_teacher_schedule")
def get_teacher_schedule(
    date_from: str,
    date_to: str,
    user: dict = require_role("teacher")
):
    """Получить расписание учителя на диапазон дат с уроками"""
    sb = get_supabase()
    
    # Получаем все уроки учителя из расписания.
    timetable = _timetable_entries_for_teacher(
        sb,
        user["id"],
        "id,class_id,class_ids,subject,weekday,start_time,end_time,room,subject_id,subjects(name)",
    )
    
    if not timetable:
        return {"lessons": []}
    
    # Собираем все class_ids для batch запроса названий
    all_cids: list[str] = []
    for e in timetable:
        cids = e.get("class_ids") or []
        if cids:
            all_cids.extend([str(c) for c in cids if c])
        elif e.get("class_id"):
            all_cids.append(str(e["class_id"]))
    all_cids = list(dict.fromkeys(all_cids))
    class_names: dict[str, str] = {}
    if all_cids:
        cn_rows = sb.table("classes").select("id,name").in_("id", all_cids).execute().data or []
        class_names = {str(r["id"]): r.get("name", "") for r in cn_rows if r.get("id")}
    
    # Генерируем даты в диапазоне
    start_date = date.fromisoformat(date_from)
    end_date = date.fromisoformat(date_to)
    
    lessons = []
    current_date = start_date
    
    while current_date <= end_date:
        wd = current_date.weekday()  # 0=Mon, 6=Sun (matches timetable_entries)
        
        # Находим уроки на этот день недели
        day_lessons = [e for e in timetable if e.get("weekday") == wd]
        
        for entry in day_lessons:
            subject_name = entry.get("subjects", {}).get("name") if entry.get("subjects") else entry.get("subject")
            cids = entry.get("class_ids") or []
            if not cids:
                cids = [entry.get("class_id")] if entry.get("class_id") else []
            
            # Для каждой группы создаём отдельный урок
            for cid in cids:
                cid_str = str(cid) if cid else ""
                lessons.append({
                    "timetable_entry_id": entry.get("id"),
                    "date": current_date.isoformat(),
                    "weekday": wd,
                    "start_time": entry.get("start_time"),
                    "end_time": entry.get("end_time"),
                    "subject": entry.get("subject"),
                    "subject_name": subject_name,
                    "class_id": cid_str,
                    "class_name": class_names.get(cid_str, ""),
                    "room": entry.get("room")
                })
        
        current_date += timedelta(days=1)
    
    # Сортируем по дате и времени
    lessons.sort(key=lambda x: (x["date"], x["start_time"] or ""))
    
    return {"lessons": lessons}


@router.get("/teacher/lessons/{lesson_date}")
@timed("get_teacher_lessons_for_date")
def get_teacher_lessons_for_date(
    lesson_date: str,
    user: dict = require_role("teacher")
):
    """Получить все уроки учителя на конкретную дату"""
    sb = get_supabase()
    
    # Определяем день недели (0=Mon, 6=Sun — совпадает с timetable_entries)
    lesson_date_obj = date.fromisoformat(lesson_date)
    wd = lesson_date_obj.weekday()
    
    # Получаем расписание на этот день.
    timetable = _timetable_entries_for_teacher(
        sb,
        user["id"],
        "id,class_id,class_ids,subject,weekday,start_time,end_time,room,subject_id,subjects(name)",
        weekday=wd,
    )
    timetable.sort(key=lambda x: (x.get("start_time") or ""))
    
    if not timetable:
        return {"lessons": []}

    # Собираем все class_ids для batch запроса названий
    all_cids: list[str] = []
    for e in timetable:
        cids = e.get("class_ids") or []
        if cids:
            all_cids.extend([str(c) for c in cids if c])
        elif e.get("class_id"):
            all_cids.append(str(e["class_id"]))
    all_cids = list(dict.fromkeys(all_cids))
    class_names: dict[str, str] = {}
    if all_cids:
        cn_rows = sb.table("classes").select("id,name").in_("id", all_cids).execute().data or []
        class_names = {str(r["id"]): r.get("name", "") for r in cn_rows if r.get("id")}

    # Оптимизация: получаем все записи журнала одним запросом
    entry_ids = [e["id"] for e in timetable]
    journal_entries = (
        sb.table("lesson_journal")
        .select("timetable_entry_id")
        .in_("timetable_entry_id", entry_ids)
        .eq("lesson_date", lesson_date)
        .execute()
        .data
        or []
    )
    
    # Создаем множество ID уроков, для которых есть записи в журнале
    entries_with_journal = {e["timetable_entry_id"] for e in journal_entries}
    
    lessons = []
    for entry in timetable:
        subject_name = entry.get("subjects", {}).get("name") if entry.get("subjects") else entry.get("subject")
        cids = entry.get("class_ids") or []
        if not cids:
            cids = [entry.get("class_id")] if entry.get("class_id") else []
        
        # Для каждой группы создаём отдельную запись урока
        for cid in cids:
            cid_str = str(cid) if cid else ""
            lessons.append({
                "timetable_entry_id": entry.get("id"),
                "date": lesson_date,
                "start_time": entry.get("start_time"),
                "end_time": entry.get("end_time"),
                "subject": entry.get("subject"),
                "subject_name": subject_name,
                "class_id": cid_str,
                "class_name": class_names.get(cid_str, ""),
                "room": entry.get("room"),
                "has_journal_entries": entry.get("id") in entries_with_journal
            })
    
    return {"lessons": lessons}


class BulkAttendanceIn(BaseModel):
    timetable_entry_id: str
    lesson_date: str
    student_ids: list[str]
    present: bool


@router.post("/bulk-attendance")
def bulk_mark_attendance(
    payload: BulkAttendanceIn,
    user: dict = require_role("teacher", "admin", "manager")
):
    """Массовая отметка посещаемости"""
    sb = get_supabase()
    
    # Проверяем урок
    entry = (
        sb.table("timetable_entries")
        .select("id,class_id,teacher_id")
        .eq("id", payload.timetable_entry_id)
        .limit(1)
        .execute()
        .data
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    # Allow teacher to edit if they are the teacher OR if they have access to the class?
    # For now, strict check on teacher_id or being admin.
    # If "All groups" mode is on, we might want to relax this?
    # But usually you only grade your own lessons. Confirmed by user "Choice of subject".
    # Assuming if I pick a subject I teach, I can grade.
    
    entry_data = entry[0]
    
    # Relaxed check: if I am a teacher, I can grade any lesson? 
    # Or should I check if I am listed as teacher_id?
    # Existing logic:
    if user["role"] == "teacher" and not _teacher_can_access_entry(sb, user["id"], entry_data):
        # Allow if no teacher assigned?
        if not entry_data.get("teacher_id"):
            pass  # Allowed
        else:
            raise HTTPException(status_code=403, detail="Access denied")
    
    # Создаем/обновляем записи для всех студентов
    records = []
    for student_id in payload.student_ids:
        records.append({
            "timetable_entry_id": payload.timetable_entry_id,
            "lesson_date": payload.lesson_date,
            "student_id": student_id,
            "present": payload.present,
            "created_by": user["id"],
            "updated_at": datetime.utcnow().isoformat()
        })
    
    if records:
        sb.table("lesson_journal").upsert(
            records,
            on_conflict="timetable_entry_id,lesson_date,student_id"
        ).execute()
    
    return {"ok": True, "updated": len(records)}


@router.get("/lesson-details")
def get_lesson_details(
    timetable_entry_id: str,
    lesson_date: str,
    user: dict = require_role("teacher", "admin", "manager")
):
    """Получить детальную информацию об уроке со списком студентов и их оценками"""
    sb = get_supabase()
    
    # Получаем информацию об уроке
    entry = (
        sb.table("timetable_entries")
        .select("id,class_id,class_ids,subject,teacher_id,start_time,end_time,room,subject_id,subjects(name)")
        .eq("id", timetable_entry_id)
        .limit(1)
        .execute()
        .data
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    entry_data = entry[0]
    
    # Проверяем доступ для учителя
    if user["role"] == "teacher" and not _teacher_can_access_entry(sb, user["id"], entry_data):
        raise HTTPException(status_code=403, detail="Access denied")

    # Поддержка multi-group: используем class_ids (массив) вместо class_id (одного)
    class_ids_list = entry_data.get("class_ids") or []
    if not class_ids_list:
        class_ids_list = [entry_data.get("class_id")] if entry_data.get("class_id") else []
    class_ids_list = [str(c) for c in class_ids_list if c]
    
    subject_name = entry_data.get("subjects", {}).get("name") if entry_data.get("subjects") else entry_data.get("subject")
    
    # Получаем названия всех классов
    class_name = ""
    if class_ids_list:
        cn_rows = sb.table("classes").select("id,name").in_("id", class_ids_list).execute().data or []
        cn_map = {str(r["id"]): r.get("name", "") for r in cn_rows if r.get("id")}
        class_name = ", ".join(cn_map.get(c, c) for c in class_ids_list)
    
    # Получаем студентов из ВСЕХ классов (для multi-group уроков)
    enrollments = []
    for cid in class_ids_list:
        enr = (
            sb.table("class_enrollments")
            .select("id,legacy_student_id,student_number,student_full_name,class_id")
            .eq("class_id", cid)
            .execute()
            .data
            or []
        )
        enrollments.extend(enr)
    
    student_ids = [e.get("legacy_student_id") for e in enrollments if e.get("legacy_student_id")]
    users_by_id: dict[str, dict] = {}
    
    if student_ids:
        students_resp = (
            sb.table("users")
            .select("id,username,full_name")
            .in_("id", student_ids)
            .execute()
            .data
            or []
        )
        users_by_id = {str(s.get("id")): s for s in students_resp if s.get("id")}
    
    # Получаем записи из журнала для этого урока
    makeup_supported = _lesson_journal_makeup_supported(sb)
    journal_select = "student_id,grade,present,comment,lesson_topic,homework,attendance_type,subject_topic_id"
    if makeup_supported:
        journal_select = (
            "student_id,grade,makeup_grade,present,comment,lesson_topic,homework,"
            "attendance_type,attendance_makeup,subject_topic_id"
        )
    journal_records = (
        sb.table("lesson_journal")
        .select(journal_select)
        .eq("timetable_entry_id", timetable_entry_id)
        .eq("lesson_date", lesson_date)
        .execute()
        .data
        or []
    )
    
    # Индексируем записи по student_id
    journal_by_student = {r.get("student_id"): r for r in journal_records}
    
    # Собираем тему и ДЗ (берем первое непустое значение)
    lesson_topic = None
    homework = None
    subject_topic_id = None
    for record in journal_records:
        if record.get("lesson_topic") and not lesson_topic:
            lesson_topic = record["lesson_topic"]
        if record.get("homework") and not homework:
            homework = record["homework"]
        if record.get("subject_topic_id") and not subject_topic_id:
            subject_topic_id = record["subject_topic_id"]
    
    # Формируем список студентов
    students = []
    for e in enrollments:
        legacy_id = e.get("legacy_student_id")
        enrollment_id = e.get("id")
        sid = str(legacy_id) if legacy_id else (str(enrollment_id) if enrollment_id else None)
        if not sid:
            continue

        user_row = users_by_id.get(str(legacy_id)) if legacy_id else None
        journal = journal_by_student.get(str(legacy_id), {}) if legacy_id else {}
        student_number = e.get("student_number")
        fallback_name = e.get("student_full_name") or (f"Student #{student_number}" if student_number is not None else "Student")
        display_name = (user_row.get("full_name") or user_row.get("username")) if user_row else fallback_name

        students.append({
            "id": sid,
            "name": display_name,
            "username": user_row.get("username") if user_row else None,
            "student_number": student_number,
            "grade": journal.get("grade"),
            "makeup_grade": journal.get("makeup_grade"),
            "present": journal.get("present"),
            "comment": journal.get("comment"),
            "attendance_type": journal.get("attendance_type"),
            "attendance_makeup": bool(journal.get("attendance_makeup")),
        })
    
    # Сортируем по номеру студента
    students.sort(key=lambda x: (x["student_number"] is None, x["student_number"] or 0, x["name"] or "", x["username"] or ""))
    
    return {
        "lesson": {
            "timetable_entry_id": timetable_entry_id,
            "date": lesson_date,
            "subject": entry_data.get("subject"),
            "subject_name": subject_name,
            "class_id": class_ids_list[0] if class_ids_list else None,
            "class_name": class_name,
            "start_time": entry_data.get("start_time"),
            "end_time": entry_data.get("end_time"),
            "room": entry_data.get("room"),
            "lesson_topic": lesson_topic,
            "homework": homework,
            "subject_topic_id": subject_topic_id
        },
        "students": students
    }


@router.get("/classes/{class_id}/journal")
def get_class_journal(
    class_id: str,
    subject_id: str | None = None,
    user: dict = require_role("teacher", "admin", "manager")
):
    """Return class journal grid: students x lesson dates."""
    sb = get_supabase()

    timetable = _timetable_entries_for_class(
        sb,
        class_id,
        "id,class_id,class_ids,subject,weekday,start_time,subject_id,teacher_id,stream_id,created_at,lesson_date",
        user["id"] if user["role"] == "teacher" else None,
    )

    # Filter by subject ID; fallback to subject name matching for legacy rows.
    if subject_id:
        target_subject_id = str(subject_id)
        subj_data = sb.table("subjects").select("name").eq("id", subject_id).limit(1).execute().data
        target_name = str(subj_data[0]["name"]) if subj_data else ""
        target_name_norm = target_name.strip().casefold()

        filtered = []
        for t in timetable:
            if t.get("subject_id") and str(t.get("subject_id")) == target_subject_id:
                filtered.append(t)
                continue
            t_name_norm = str(t.get("subject") or "").strip().casefold()
            if target_name_norm and (
                t_name_norm == target_name_norm
                or target_name_norm in t_name_norm
                or t_name_norm in target_name_norm
            ):
                filtered.append(t)
                continue
        
        timetable = filtered

    if not timetable:
        return {"students": [], "lessons": [], "grades": {}}

    enrollments = (
        sb.table("class_enrollments")
        .select("id,legacy_student_id,student_full_name,student_number")
        .eq("class_id", class_id)
        .execute()
        .data
        or []
    )
    student_ids = [e.get("legacy_student_id") for e in enrollments if e.get("legacy_student_id")]

    users_by_id: dict[str, dict] = {}
    if student_ids:
        user_rows = (
            sb.table("users")
            .select("id,username,full_name")
            .in_("id", student_ids)
            .execute()
            .data
            or []
        )
        for u in user_rows:
            users_by_id[str(u["id"])] = u

    # Build students list from enrollments, merging user data where available
    students = []
    for enr in enrollments:
        lid = enr.get("legacy_student_id")
        user_data = users_by_id.get(str(lid)) if lid else None
        if user_data:
            students.append(user_data)
        else:
            # Student without user account - use enrollment data directly
            students.append({
                "id": str(enr.get("id")),
                "full_name": enr.get("student_full_name"),
                "username": None,
                "student_number": enr.get("student_number"),
            })
    students.sort(key=lambda s: (s.get("student_number") is None, s.get("student_number") or 0, s.get("full_name") or ""))

    entry_ids = [e.get("id") for e in timetable if e.get("id")]
    journal_records = []
    makeup_supported = _lesson_journal_makeup_supported(sb)
    if entry_ids:
        journal_select = (
            "timetable_entry_id,lesson_date,student_id,grade,present,comment,"
            "lesson_topic,homework,attendance_type,subject_topic_id,created_by,updated_at"
        )
        if makeup_supported:
            journal_select = (
                "timetable_entry_id,lesson_date,student_id,grade,makeup_grade,present,comment,"
                "lesson_topic,homework,attendance_type,attendance_makeup,subject_topic_id,created_by,updated_at"
            )
        journal_records = (
            sb.table("lesson_journal")
            .select(journal_select)
            .in_("timetable_entry_id", entry_ids)
            .order("lesson_date", desc=False)
            .execute()
            .data
            or []
        )

    creator_ids = sorted({str(r.get("created_by")) for r in journal_records if r.get("created_by")})
    creator_names: dict[str, str] = {}
    if creator_ids:
        creator_rows = (
            sb.table("users")
            .select("id,full_name,username")
            .in_("id", creator_ids)
            .execute()
            .data
            or []
        )
        for row in creator_rows:
            uid = row.get("id")
            if not uid:
                continue
            creator_names[str(uid)] = str(row.get("full_name") or row.get("username") or "Unknown")

    entry_by_id = {str(e.get("id")): e for e in timetable if e.get("id")}
    lessons_by_date: dict[str, dict] = {}

    today = date.today()
    journal_future_days = 7
    if today.month >= 9:
        start_date = date(today.year, 9, 1)
    else:
        start_date = date(today.year - 1, 9, 1)

    end_date = today + timedelta(days=journal_future_days)

    current = start_date
    while current <= end_date:
        wd = current.weekday()
        d_str = current.isoformat()

        day_entries = []
        for e in timetable:
            lesson_date = _normalize_lesson_date(e.get("lesson_date"))
            if lesson_date:
                if lesson_date == d_str:
                    day_entries.append(e)
            else:
                if _entry_matches_weekday(e.get("weekday"), wd):
                    day_entries.append(e)

        for entry in day_entries:
            key = f"{d_str}_{entry['id']}"
            lessons_by_date[key] = {
                "date": d_str,
                "timetable_entry_id": entry['id'],
                "subject_name": entry.get("subject", ""),
                "start_time": entry.get("start_time"),
                "lesson_topic": None,
                "homework": None
            }
        current += timedelta(days=1)

    lesson_info: dict[str, dict[str, str | None]] = {}
    records_by_cell: dict[tuple[str, str, str], list[dict]] = {}

    for record in journal_records:
        entry_id = record.get("timetable_entry_id")
        d_str = record.get("lesson_date")
        if not entry_id or not d_str:
            continue

        entry = entry_by_id.get(str(entry_id))
        if not entry:
            continue

        key = f"{d_str}_{entry_id}"
        if key not in lessons_by_date:
            lessons_by_date[key] = {
                "date": d_str,
                "timetable_entry_id": entry_id,
                "subject_name": entry.get("subject", ""),
                "start_time": entry.get("start_time"),
                "lesson_topic": None,
                "homework": None,
                "subject_topic_id": None
            }

        info = lesson_info.setdefault(key, {"lesson_topic": None, "homework": None, "subject_topic_id": None})
        if record.get("lesson_topic") and not info["lesson_topic"]:
            info["lesson_topic"] = record["lesson_topic"]
        if record.get("homework") and not info["homework"]:
            info["homework"] = record["homework"]
        if record.get("subject_topic_id") and not info["subject_topic_id"]:
            info["subject_topic_id"] = record["subject_topic_id"]

        student_id = record.get("student_id")
        if student_id:
            cell = (str(student_id), str(entry_id), str(d_str))
            records_by_cell.setdefault(cell, []).append(record)

    lessons = sorted(
        lessons_by_date.values(),
        key=lambda x: (
            x.get("date") or "",
            str(x.get("start_time") or ""),
            str(x.get("timetable_entry_id") or ""),
        ),
    )
    for lesson in lessons:
        key = f"{lesson['date']}_{lesson['timetable_entry_id']}"
        info = lesson_info.get(key, {})
        if info.get("lesson_topic"):
            lesson["lesson_topic"] = info.get("lesson_topic")
        if info.get("homework"):
            lesson["homework"] = info.get("homework")
        if info.get("subject_topic_id"):
            lesson["subject_topic_id"] = info.get("subject_topic_id")

    grades = {}
    for student in students:
        sid = student.get("id")
        grades[sid] = {}
        for lesson in lessons:
            key = f"{lesson['date']}_{lesson['timetable_entry_id']}"

            student_records = records_by_cell.get((str(sid), str(lesson["timetable_entry_id"]), str(lesson["date"])), [])
            sorted_records = sorted(
                student_records,
                key=lambda r: str(r.get("updated_at") or ""),
                reverse=True,
            )
            latest_record = sorted_records[0] if sorted_records else None

            lesson_grades = [
                {
                    "grade": r.get("grade"),
                    "comment": r.get("comment"),
                    "created_by": r.get("created_by"),
                    "created_by_name": creator_names.get(str(r.get("created_by")), "Unknown")
                    if r.get("created_by")
                    else "Unknown"
                }
                for r in sorted_records
                if r.get("grade") is not None
            ]

            present = latest_record.get("present") if latest_record else None
            attendance_type = latest_record.get("attendance_type") if latest_record else None
            makeup_grade = latest_record.get("makeup_grade") if latest_record else None
            attendance_makeup = bool(latest_record.get("attendance_makeup")) if latest_record else False
            marked_by_name = (
                creator_names.get(str(latest_record.get("created_by")), "Unknown")
                if latest_record and latest_record.get("created_by")
                else None
            )
                    
            grades[sid][key] = {
                "grades": lesson_grades,
                "present": present,
                "attendance_type": attendance_type,
                "makeup_grade": makeup_grade,
                "attendance_makeup": attendance_makeup,
                "marked_by_name": marked_by_name
            }

    students_list = [
        {
            "id": s.get("id"),
            "name": s.get("full_name") or s.get("username") or "—",
            "username": s.get("username"),
            "student_number": s.get("student_number"),
        }
        for s in students
    ]

    return {
        "students": students_list,
        "lessons": lessons,
        "grades": grades
    }


@router.post("/classes/{class_id}/grades")
def add_grade(
    class_id: str,
    payload: AddGradeIn,
    user: dict = require_role("teacher", "admin", "manager")
):
    """Partial create/update for lesson grade/attendance/comment."""
    sb = get_supabase()
    makeup_supported = _lesson_journal_makeup_supported(sb)

    fields_set = set(payload.model_fields_set)
    mutating_fields = {"grade", "makeup_grade", "present", "comment", "attendance_type", "attendance_makeup"}
    if fields_set.isdisjoint(mutating_fields):
        raise HTTPException(status_code=400, detail="No grade/attendance fields to update")
    if not makeup_supported and {"makeup_grade", "attendance_makeup"} & fields_set:
        raise HTTPException(
            status_code=400,
            detail="DB schema is missing makeup fields. Apply migration supabase/migrations/20260305_000001_lesson_journal_makeup.sql and restart the API.",
        )

    if "grade" in fields_set and payload.grade is not None and (payload.grade < 2 or payload.grade > 5):
        raise HTTPException(status_code=400, detail="Grade must be between 2 and 5")
    if "makeup_grade" in fields_set and payload.makeup_grade is not None and (payload.makeup_grade < 2 or payload.makeup_grade > 5):
        raise HTTPException(status_code=400, detail="Makeup grade must be between 2 and 5")

    valid_attendance_types = ["present", "absent", "duty", "excused", "sick", None]
    if "attendance_type" in fields_set and payload.attendance_type not in valid_attendance_types:
        raise HTTPException(status_code=400, detail=f"Invalid attendance_type. Must be one of: {valid_attendance_types}")

    entry = (
        sb.table("timetable_entries")
        .select("id,class_id,class_ids,teacher_id,subject_id")
        .eq("id", payload.timetable_entry_id)
        .limit(1)
        .execute()
        .data
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Lesson not found")

    entry_data = entry[0]
    entry_class_ids = entry_data.get("class_ids") or []
    if class_id not in entry_class_ids and entry_data.get("class_id") != class_id:
        raise HTTPException(status_code=400, detail="Lesson does not belong to this class")

    if user["role"] == "teacher" and not _teacher_can_access_entry(sb, user["id"], entry_data):
        raise HTTPException(status_code=403, detail="Access denied")

    resolved_student_id = _resolve_legacy_student_id_for_class(sb, class_id, payload.student_id)
    if not resolved_student_id:
        raise HTTPException(status_code=404, detail="Student not in this class")

    d = date.fromisoformat(payload.lesson_date)

    existing = (
        sb.table("lesson_journal")
        .select("timetable_entry_id")
        .eq("timetable_entry_id", payload.timetable_entry_id)
        .eq("lesson_date", d.isoformat())
        .eq("student_id", resolved_student_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    is_update = bool(existing)

    record_data = {
        "timetable_entry_id": payload.timetable_entry_id,
        "lesson_date": d.isoformat(),
        "student_id": resolved_student_id,
        "created_by": user["id"],
        "updated_at": datetime.utcnow().isoformat(),
    }

    if "grade" in fields_set:
        record_data["grade"] = payload.grade
    if "makeup_grade" in fields_set:
        record_data["makeup_grade"] = payload.makeup_grade
    if "comment" in fields_set:
        record_data["comment"] = payload.comment

    if "attendance_type" in fields_set:
        record_data["attendance_type"] = payload.attendance_type
        if payload.attendance_type is not None:
            record_data["present"] = payload.attendance_type in ["present", "duty"]
        elif "present" in fields_set:
            record_data["present"] = bool(payload.present) if payload.present is not None else None
        if makeup_supported and "attendance_makeup" not in fields_set and payload.attendance_type not in ABSENCE_ATTENDANCE_TYPES:
            record_data["attendance_makeup"] = False
    elif "present" in fields_set:
        record_data["present"] = bool(payload.present) if payload.present is not None else None

    if makeup_supported and "attendance_makeup" in fields_set:
        record_data["attendance_makeup"] = bool(payload.attendance_makeup)

    if not is_update:
        record_data.setdefault("grade", None)
        record_data.setdefault("present", True)
        record_data.setdefault("comment", None)
        record_data.setdefault("attendance_type", None)
        if makeup_supported:
            record_data.setdefault("makeup_grade", None)
            record_data.setdefault("attendance_makeup", False)

    sb.table("lesson_journal").upsert(
        record_data,
        on_conflict="timetable_entry_id,lesson_date,student_id",
    ).execute()

    return {"ok": True}


@router.delete("/grades/{grade_id}")
def delete_grade(grade_id: str, user: dict = require_role("teacher", "admin", "manager")):
    """Удалить оценку"""
    sb = get_supabase()
    
    # Получаем оценку для проверки доступа
    grade = (
        sb.table("subject_grades")
        .select("class_id,subject,created_by")
        .eq("id", grade_id)
        .limit(1)
        .execute()
        .data
    )
    
    if not grade:
        raise HTTPException(status_code=404, detail="Grade not found")
    
    grade_data = grade[0]
    
    # Проверяем доступ
    if user["role"] == "teacher":
        # Учитель может удалить только свои оценки по своим предметам
        # Check both class_ids array and legacy class_id field
        g_class_id = grade_data.get("class_id")
        timetable = (
            sb.table("timetable_entries")
            .select("id")
            .cs("class_ids", [g_class_id])
            .eq("teacher_id", user["id"])
            .eq("subject", grade_data.get("subject"))
            .limit(1)
            .execute()
            .data
        )
        if not timetable:
            timetable = (
                sb.table("timetable_entries")
                .select("id")
                .eq("class_id", g_class_id)
                .eq("teacher_id", user["id"])
                .eq("subject", grade_data.get("subject"))
                .limit(1)
                .execute()
                .data
            )
        if not timetable or grade_data.get("created_by") != user["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
    
    sb.table("subject_grades").delete().eq("id", grade_id).execute()
    
    return {"ok": True}


@router.get("/classes/{class_id}/export")
def export_journal(class_id: str, user: dict = require_role("teacher", "admin", "manager")):
    """Экспорт журнала в Excel"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    sb = get_supabase()
    
    # Получаем данные журнала
    journal = get_class_journal(class_id=class_id, user=user)
    students = journal["students"]
    lessons = journal["lessons"]
    grades_data = journal["grades"]
    
    # Получаем название класса
    cls = sb.table("classes").select("name").eq("id", class_id).limit(1).execute().data
    class_name = cls[0].get("name") if cls else "Class"
    
    # Создаем Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=500, detail="Failed to create worksheet")
    ws.title = "Journal"
    
    # Заголовки
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws.cell(1, 1, "Student").fill = header_fill
    ws.cell(1, 1).font = header_font
    
    # Даты уроков
    for col_idx, lesson in enumerate(lessons, start=2):
        cell = ws.cell(1, col_idx, f"{lesson['date']}\n{lesson['subject_name']}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Средний балл
    avg_col = len(lessons) + 2
    cell = ws.cell(1, avg_col, "Average")
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Данные
    for row_idx, student in enumerate(students, start=2):
        ws.cell(row_idx, 1, student["name"])
        
        sid = student["id"]
        all_grades = []
        
        for col_idx, lesson in enumerate(lessons, start=2):
            key = f"{lesson['date']}_{lesson['timetable_entry_id']}"
            cell_data = grades_data.get(sid, {}).get(key, {})
            student_grades = cell_data.get("grades", []) if isinstance(cell_data, dict) else []
            if student_grades:
                # Берем все оценки
                grade_values = [g["grade"] for g in student_grades]
                # Показываем все оценки через запятую
                cell = ws.cell(row_idx, col_idx, ", ".join(map(str, grade_values)))
                all_grades.extend(grade_values)
        
        # Средний балл
        if all_grades:
            avg = sum(all_grades) / len(all_grades)
            ws.cell(row_idx, avg_col, round(avg, 2))
    
    # Ширина колонок
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, len(lessons) + 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Сохраняем
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{class_name}_zhurnal.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


class UpdateLessonInfoIn(BaseModel):
    timetable_entry_id: str
    lesson_date: str  # YYYY-MM-DD
    lesson_topic: str | None = None
    homework: str | None = None
    subject_topic_id: str | None = None


@router.post("/classes/{class_id}/lesson-info")
def update_lesson_info(
    class_id: str,
    payload: UpdateLessonInfoIn,
    user: dict = require_role("teacher", "admin", "manager")
):
    """Обновить тему урока и домашнее задание"""
    sb = get_supabase()
    
    # Проверяем урок
    entry = (
        sb.table("timetable_entries")
        .select("id,class_id,class_ids,teacher_id,subject_id")
        .eq("id", payload.timetable_entry_id)
        .limit(1)
        .execute()
        .data
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    entry_data = entry[0]
    # Verify lesson belongs to this class (class_ids array or legacy class_id)
    entry_class_ids = entry_data.get("class_ids") or []
    if class_id not in entry_class_ids and entry_data.get("class_id") != class_id:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    
    
    # Проверяем права
    if user["role"] == "teacher" and not _teacher_can_access_entry(sb, user["id"], entry_data):
        raise HTTPException(status_code=403, detail="Not your lesson")
    
    # Проверяем есть ли уже записи для этого урока/даты
    existing = (
        sb.table("lesson_journal")
        .select("*")
        .eq("timetable_entry_id", payload.timetable_entry_id)
        .eq("lesson_date", payload.lesson_date)
        .execute()
        .data or []
    )
    
    update_data = {}
    if "lesson_topic" in payload.model_fields_set:
        update_data["lesson_topic"] = payload.lesson_topic
    if "homework" in payload.model_fields_set:
        update_data["homework"] = payload.homework
    if "subject_topic_id" in payload.model_fields_set:
        update_data["subject_topic_id"] = payload.subject_topic_id
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    if existing:
        # Обновляем все существующие записи для этого урока
        sb.table("lesson_journal").update(update_data).eq("timetable_entry_id", payload.timetable_entry_id).eq("lesson_date", payload.lesson_date).execute()
    else:
        # If there are no rows yet, create rows for all class enrollments.
        # Provision missing legacy_student_id to keep lesson_journal consistent.
        enrollments_resp = (
            sb.table("class_enrollments")
            .select("id,legacy_student_id,student_full_name,student_number")
            .eq("class_id", class_id)
            .execute()
        )
        enrollments = enrollments_resp.data or []
        
        if enrollments:
            insert_records = []
            for enr in enrollments:
                student_id = enr.get("legacy_student_id")
                if not student_id:
                    student_id = _ensure_legacy_student_id_for_enrollment(sb, class_id, enr)
                if not student_id:
                    continue
                insert_records.append(
                    {
                        "timetable_entry_id": payload.timetable_entry_id,
                        "lesson_date": payload.lesson_date,
                        "student_id": student_id,
                        "created_by": user["id"],
                        **update_data,
                    }
                )
            if insert_records:
                sb.table("lesson_journal").insert(insert_records).execute()
    
    return {"success": True, "message": "Lesson info updated"}


@router.get("/classes/{class_id}/lesson-info")
def get_lesson_info(
    class_id: str,
    timetable_entry_id: str,
    lesson_date: str,
    user: dict = Depends(get_current_user),
):
    """Получить тему урока и домашнее задание для конкретного урока"""
    sb = get_supabase()
    
    # Проверяем урок
    entry = (
        sb.table("timetable_entries")
        .select("id,class_id,class_ids,subject,teacher_id,subject_id,subjects(name)")
        .eq("id", timetable_entry_id)
        .limit(1)
        .execute()
        .data
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Lesson not found")
    
    entry_data = entry[0]
    # Verify lesson belongs to this class (class_ids array or legacy class_id)
    entry_class_ids = entry_data.get("class_ids") or []
    if class_id not in entry_class_ids and entry_data.get("class_id") != class_id:
        raise HTTPException(status_code=404, detail="Lesson not found")
    subject_name = entry_data.get("subjects", {}).get("name") if entry_data.get("subjects") else entry_data.get("subject")
    
    # Проверяем доступ
    if user["role"] == "teacher" and not _teacher_can_access_entry(sb, user["id"], entry_data):
        raise HTTPException(status_code=403, detail="Not your lesson")
    
    if user["role"] == "student":
        # Студент может видеть только если он в этом классе
        enrollment = (
            sb.table("class_enrollments")
            .select("id")
            .eq("class_id", class_id)
            .eq("student_id", user["id"])
            .limit(1)
            .execute()
            .data
        )
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not in this class")
    
    # Получаем тему и ДЗ
    records = (
        sb.table("lesson_journal")
        .select("lesson_topic,homework,subject_topic_id")
        .eq("timetable_entry_id", timetable_entry_id)
        .eq("lesson_date", lesson_date)
        .execute()
        .data
        or []
    )
    
    # Берем первую запись с непустыми полями
    lesson_topic = None
    homework = None
    subject_topic_id = None
    
    for record in records:
        if record.get("lesson_topic") and not lesson_topic:
            lesson_topic = record["lesson_topic"]
        if record.get("homework") and not homework:
            homework = record["homework"]
        if record.get("subject_topic_id") and not subject_topic_id:
            subject_topic_id = record["subject_topic_id"]
    
    return {
        "lesson_topic": lesson_topic,
        "homework": homework,
        "subject_topic_id": subject_topic_id,
        "subject": entry_data.get("subject"),
        "subject_name": subject_name,
        "lesson_date": lesson_date
    }


@router.get("/student/homework")
def get_student_homework(user: dict = require_role("student")):
    """Получить все домашние задания для студента"""
    sb = get_supabase()
    
    # Получаем классы студента
    enrollments = (
        sb.table("class_enrollments")
        .select("class_id")
        .eq("student_id", user["id"])
        .execute()
        .data
        or []
    )
    
    class_ids = [e.get("class_id") for e in enrollments if e.get("class_id")]
    if not class_ids:
        return {"homework": []}
    
    # Получаем расписание этих классов
    timetable = (
        sb.table("timetable_entries")
        .select("id,class_id,subject,weekday,start_time,subject_id,classes(name),subjects(name)")
        .in_("class_id", class_ids)
        .execute()
        .data
        or []
    )
    
    timetable_ids = [e.get("id") for e in timetable]
    if not timetable_ids:
        return {"homework": []}
    
    # Получаем все ДЗ за последние 30 дней
    from datetime import date, timedelta
    date_from = (date.today() - timedelta(days=30)).isoformat()
    
    homework_records = (
        sb.table("lesson_journal")
        .select("timetable_entry_id,lesson_date,homework,lesson_topic")
        .in_("timetable_entry_id", timetable_ids)
        .gte("lesson_date", date_from)
        .not_.is_("homework", "null")
        .order("lesson_date", desc=True)
        .execute()
        .data
        or []
    )
    
    # Группируем и форматируем
    entry_by_id = {str(e.get("id")): e for e in timetable if e.get("id")}
    result = []
    for record in homework_records:
        entry_id = record.get("timetable_entry_id")
        entry = entry_by_id.get(str(entry_id))
        if not entry:
            continue
        
        class_name = entry.get("classes", {}).get("name", "Unknown") if entry.get("classes") else "Unknown"
        subject_name = entry.get("subjects", {}).get("name") if entry.get("subjects") else entry.get("subject", "Unknown")
        
        result.append({
            "lesson_date": record.get("lesson_date"),
            "subject": entry.get("subject", "Unknown"),
            "subject_name": subject_name,
            "class_name": class_name,
            "lesson_topic": record.get("lesson_topic"),
            "homework": record.get("homework"),
            "timetable_entry_id": entry_id
        })
    
    return {"homework": result}
