from __future__ import annotations

from datetime import date, datetime
from collections import defaultdict
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.core.deps import get_current_user, require_role
from app.db.supabase_client import get_supabase

router = APIRouter()


def _lesson_journal_supported(sb) -> bool:
    try:
        sb.table("lesson_journal").select("timetable_entry_id").limit(1).execute()
        return True
    except Exception:
        return False


class CreateAssessmentIn(BaseModel):
    class_id: str
    title: str
    date: str  # YYYY-MM-DD


@router.post("/assessments")
def create_assessment(payload: CreateAssessmentIn, user: dict = require_role("teacher", "admin")):
    sb = get_supabase()
    resp = sb.table("assessments").insert({**payload.model_dump(), "created_by": user["id"]}).execute()
    return {"assessment": resp.data[0] if isinstance(resp.data, list) and resp.data else resp.data}


@router.get("/classes/{class_id}/assessments")
def list_assessments(class_id: str, _user: dict = Depends(get_current_user)):
    sb = get_supabase()
    resp = sb.table("assessments").select("*").eq("class_id", class_id).order("date", desc=True).execute()
    return {"assessments": resp.data or []}


class GradeUpsert(BaseModel):
    student_id: str
    value: int  # 1..5
    comment: str | None = None


class SetGradesIn(BaseModel):
    grades: list[GradeUpsert]


@router.put("/assessments/{assessment_id}/grades")
def set_grades(assessment_id: str, payload: SetGradesIn, _: dict = require_role("teacher", "admin")):
    sb = get_supabase()
    rows = [
        {
            "assessment_id": assessment_id,
            "student_id": g.student_id,
            "value": g.value,
            "comment": g.comment,
        }
        for g in payload.grades
    ]
    # upsert by composite key (assessment_id, student_id)
    sb.table("grades").upsert(rows, on_conflict="assessment_id,student_id").execute()
    return {"ok": True}


@router.get("/me")
def my_grades(user: dict = require_role("student")):
    sb = get_supabase()
    resp = (
        sb.table("grades")
        .select("value, comment, assessments(id,title,date,class_id)")
        .eq("student_id", user["id"])
        .execute()
    )

    if not _lesson_journal_supported(sb):
        return {"grades": resp.data or [], "lessonJournal": []}

    # Lesson journal (per timetable lesson occurrence)
    lj = (
        sb.table("lesson_journal")
        .select("timetable_entry_id,lesson_date,present,grade,comment,created_by")
        .eq("student_id", user["id"])
        .order("lesson_date", desc=True)
        .execute()
    )
    lj_rows = lj.data or []
    entry_ids = sorted({r.get("timetable_entry_id") for r in lj_rows if r.get("timetable_entry_id")})
    teacher_ids = sorted({r.get("created_by") for r in lj_rows if r.get("created_by")})

    entries_by_id: dict[str, dict] = {}
    if entry_ids:
        e_rows = (
            sb.table("timetable_entries")
            .select("id,class_id,teacher_id,subject,weekday,start_time,end_time,room")
            .in_("id", entry_ids)
            .execute()
            .data
            or []
        )
        entries_by_id = {e.get("id"): e for e in e_rows if e.get("id")}

    class_ids = sorted({e.get("class_id") for e in entries_by_id.values() if e.get("class_id")})
    classes_by_id: dict[str, dict] = {}
    if class_ids:
        c_rows = sb.table("classes").select("id,name").in_("id", class_ids).execute().data or []
        classes_by_id = {c.get("id"): c for c in c_rows if c.get("id")}

    teacher_by_id: dict[str, dict] = {}
    all_teacher_ids = sorted({e.get("teacher_id") for e in entries_by_id.values() if e.get("teacher_id")} | set(teacher_ids))
    if all_teacher_ids:
        t_rows = sb.table("users").select("id,full_name,username").in_("id", all_teacher_ids).execute().data or []
        teacher_by_id = {t.get("id"): t for t in t_rows if t.get("id")}

    lesson_items = []
    for r in lj_rows:
        entry = entries_by_id.get(r.get("timetable_entry_id")) or {}
        cls = classes_by_id.get(entry.get("class_id")) or {}
        tch = teacher_by_id.get(entry.get("teacher_id")) or teacher_by_id.get(r.get("created_by")) or {}
        lesson_items.append(
            {
                "timetable_entry_id": r.get("timetable_entry_id"),
                "date": r.get("lesson_date"),
                "present": r.get("present"),
                "grade": r.get("grade"),
                "comment": r.get("comment"),
                "subject": entry.get("subject"),
                "room": entry.get("room"),
                "class_id": entry.get("class_id"),
                "class_name": cls.get("name"),
                "teacher_id": entry.get("teacher_id"),
                "teacher_name": tch.get("full_name") or tch.get("username"),
                "start_time": str(entry.get("start_time"))[:5] if entry.get("start_time") else None,
                "end_time": str(entry.get("end_time"))[:5] if entry.get("end_time") else None,
            }
        )

    return {"grades": resp.data or [], "lessonJournal": lesson_items}


class LessonJournalSaveRow(BaseModel):
    student_id: str
    present: bool | None = None
    grade: int | None = None
    comment: str | None = None


@router.get("/lessons/{timetable_entry_id}")
def lesson_journal_get(
    timetable_entry_id: str,
    lesson_date: str,
    user: dict = require_role("teacher"),
):
    sb = get_supabase()
    d = date.fromisoformat(lesson_date)

    if not _lesson_journal_supported(sb):
        raise HTTPException(
            status_code=400,
            detail=(
                "DB schema is missing lesson_journal. Apply migration "
                "supabase/migrations/20251224_000004_lesson_journal.sql and restart the API."
            ),
        )

    # Ensure this is the teacher's own lesson
    e_rows = (
        sb.table("timetable_entries")
        .select("id,class_id,teacher_id,subject,room,start_time,end_time")
        .eq("id", timetable_entry_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    entry = e_rows[0] if e_rows else None
    if not entry:
        raise HTTPException(status_code=404, detail="Lesson not found")
    if entry.get("teacher_id") != user.get("id"):
        raise HTTPException(status_code=403, detail="Forbidden")

    # Load class roster
    enr = (
        sb.table("class_enrollments")
        .select("legacy_student_id")
        .eq("class_id", entry.get("class_id"))
        .execute()
        .data
        or []
    )
    student_ids = [r.get("legacy_student_id") for r in enr if r.get("legacy_student_id")]
    if not student_ids:
        return {"lesson": entry, "students": []}

    u_rows = (
        sb.table("users")
        .select("id,username,full_name")
        .in_("id", student_ids)
        .execute()
        .data
        or []
    )
    students_by_id = {s.get("id"): s for s in u_rows if s.get("id")}

    marks = (
        sb.table("lesson_journal")
        .select("student_id,present,grade,comment")
        .eq("timetable_entry_id", timetable_entry_id)
        .eq("lesson_date", d.isoformat())
        .in_("student_id", student_ids)
        .execute()
        .data
        or []
    )
    marks_by_student = {m.get("student_id"): m for m in marks if m.get("student_id")}

    out = []
    for sid in sorted(student_ids, key=lambda x: (students_by_id.get(x, {}).get("full_name") or "", students_by_id.get(x, {}).get("username") or "")):
        s = students_by_id.get(sid) or {"id": sid, "username": "", "full_name": None}
        m = marks_by_student.get(sid) or {}
        out.append(
            {
                "id": s.get("id"),
                "username": s.get("username"),
                "full_name": s.get("full_name"),
                "present": m.get("present"),
                "grade": m.get("grade"),
                "comment": m.get("comment"),
            }
        )

    return {"lesson": {**entry, "date": d.isoformat()}, "students": out}


class LessonJournalSaveIn(BaseModel):
    lesson_date: str  # YYYY-MM-DD
    rows: list[LessonJournalSaveRow]


@router.put("/lessons/{timetable_entry_id}")
def lesson_journal_save(
    timetable_entry_id: str,
    payload: LessonJournalSaveIn,
    user: dict = require_role("teacher"),
):
    sb = get_supabase()
    d = date.fromisoformat(payload.lesson_date)

    if not _lesson_journal_supported(sb):
        raise HTTPException(
            status_code=400,
            detail=(
                "DB schema is missing lesson_journal. Apply migration "
                "supabase/migrations/20251224_000004_lesson_journal.sql and restart the API."
            ),
        )

    e_rows = (
        sb.table("timetable_entries")
        .select("id,class_id,teacher_id")
        .eq("id", timetable_entry_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    entry = e_rows[0] if e_rows else None
    if not entry:
        raise HTTPException(status_code=404, detail="Lesson not found")
    if entry.get("teacher_id") != user.get("id"):
        raise HTTPException(status_code=403, detail="Forbidden")

    # Only allow writing for students in this class
    valid = (
        sb.table("class_enrollments")
        .select("legacy_student_id")
        .eq("class_id", entry.get("class_id"))
        .execute()
        .data
        or []
    )
    valid_ids = {r.get("legacy_student_id") for r in valid if r.get("legacy_student_id")}

    upserts = []
    for r in payload.rows:
        if r.student_id not in valid_ids:
            continue
        if r.grade is not None and (r.grade < 1 or r.grade > 5):
            raise HTTPException(status_code=400, detail="Grade must be 1..5")
        upserts.append(
            {
                "timetable_entry_id": timetable_entry_id,
                "lesson_date": d.isoformat(),
                "student_id": r.student_id,
                "present": r.present,
                "grade": r.grade,
                "comment": r.comment,
                "created_by": user.get("id"),
                "updated_at": datetime.utcnow().isoformat(),
            }
        )

    if upserts:
        sb.table("lesson_journal").upsert(upserts, on_conflict="timetable_entry_id,lesson_date,student_id").execute()

    return {"ok": True}


@router.get("/classes/{class_id}/journal")
def class_journal_by_dates(
    class_id: str,
    user: dict = require_role("teacher", "admin", "manager"),
):
    """
    Журнал посещаемости: таблица [ученики × даты]
    """
    sb = get_supabase()

    if not _lesson_journal_supported(sb):
        raise HTTPException(status_code=400, detail="Lesson journal not supported")

    # Получаем все уроки класса
    timetable = (
        sb.table("timetable_entries")
        .select("id,subject,teacher_id")
        .eq("class_id", class_id)
        .execute()
        .data
        or []
    )
    entry_ids = [e.get("id") for e in timetable if e.get("id")]
    if not entry_ids:
        return {"students": [], "dates": [], "data": {}}

    # Получаем все записи журнала для этого класса
    journal_records = (
        sb.table("lesson_journal")
        .select("timetable_entry_id,lesson_date,student_id,present,grade,comment")
        .in_("timetable_entry_id", entry_ids)
        .order("lesson_date", desc=False)
        .execute()
        .data
        or []
    )

    # Список учеников - напрямую из class_enrollments
    enrollments = (
        sb.table("class_enrollments")
        .select("id,student_full_name,student_number,legacy_student_id")
        .eq("class_id", class_id)
        .order("student_number")
        .execute()
        .data
        or []
    )
    
    if not enrollments:
        return {"students": [], "dates": [], "data": {}}

    # Формируем список студентов
    students = sorted(
        [
            {
                "id": e.get("id"),  # Используем enrollment ID
                "name": e.get("student_full_name") or f"Ученик #{e.get('student_number')}",
                "student_number": e.get("student_number"),
                "legacy_student_id": e.get("legacy_student_id"),
            }
            for e in enrollments
        ],
        key=lambda x: (x.get("student_number") or 999, x["name"]),
    )
    
    # Создаём маппинг legacy_student_id -> enrollment_id для журнала
    legacy_to_enrollment = {e.get("legacy_student_id"): e.get("id") for e in enrollments if e.get("legacy_student_id")}

    # Собираем все уникальные даты
    dates = sorted({r.get("lesson_date") for r in journal_records if r.get("lesson_date")})

    # Структура данных: {enrollment_id: {date: {present, grade, comment}}}
    data = defaultdict(dict)
    for rec in journal_records:
        legacy_sid = rec.get("student_id")
        enrollment_id = legacy_to_enrollment.get(legacy_sid)
        d = rec.get("lesson_date")
        if enrollment_id and d:
            data[enrollment_id][d] = {
                "present": rec.get("present"),
                "grade": rec.get("grade"),
                "comment": rec.get("comment"),
            }

    return {"students": students, "dates": dates, "data": data}


@router.get("/classes/{class_id}/journal/by-subject")
def class_journal_by_subject(
    class_id: str,
    user: dict = require_role("teacher", "admin", "manager"),
):
    """
    Журнал оценок по предметам: таблица [ученики × предметы] с средними оценками
    """
    sb = get_supabase()

    if not _lesson_journal_supported(sb):
        raise HTTPException(status_code=400, detail="Lesson journal not supported")

    # Получаем все уроки класса
    timetable = (
        sb.table("timetable_entries")
        .select("id,subject,teacher_id")
        .eq("class_id", class_id)
        .execute()
        .data
        or []
    )
    entry_ids = [e.get("id") for e in timetable if e.get("id")]
    entries_by_id = {e.get("id"): e for e in timetable if e.get("id")}

    if not entry_ids:
        return {"students": [], "subjects": [], "data": {}}

    # Получаем все записи журнала с оценками
    journal_records = (
        sb.table("lesson_journal")
        .select("timetable_entry_id,lesson_date,student_id,grade")
        .in_("timetable_entry_id", entry_ids)
        .not_.is_("grade", "null")
        .execute()
        .data
        or []
    )

    # Список учеников - напрямую из class_enrollments
    enrollments = (
        sb.table("class_enrollments")
        .select("id,student_full_name,student_number,legacy_student_id")
        .eq("class_id", class_id)
        .order("student_number")
        .execute()
        .data
        or []
    )
    
    if not enrollments:
        return {"students": [], "subjects": [], "data": {}}

    # Формируем список студентов
    students = sorted(
        [
            {
                "id": e.get("id"),  # Используем enrollment ID
                "name": e.get("student_full_name") or f"Ученик #{e.get('student_number')}",
                "student_number": e.get("student_number"),
                "legacy_student_id": e.get("legacy_student_id"),
            }
            for e in enrollments
        ],
        key=lambda x: (x.get("student_number") or 999, x["name"]),
    )
    
    # Создаём маппинг legacy_student_id -> enrollment_id для журнала
    legacy_to_enrollment = {e.get("legacy_student_id"): e.get("id") for e in enrollments if e.get("legacy_student_id")}
    enrollment_ids = [e.get("id") for e in enrollments]

    # Собираем предметы
    subjects = sorted({e.get("subject") for e in timetable if e.get("subject")})

    # Группируем оценки по ученику и предмету
    # {enrollment_id: {subject: [grades]}}
    grades_by_student_subject = defaultdict(lambda: defaultdict(list))

    for rec in journal_records:
        entry_id = rec.get("timetable_entry_id")
        entry = entries_by_id.get(entry_id)
        if not entry:
            continue

        subject = entry.get("subject")
        legacy_sid = rec.get("student_id")
        enrollment_id = legacy_to_enrollment.get(legacy_sid)
        grade = rec.get("grade")

        if enrollment_id and subject and grade:
            grades_by_student_subject[enrollment_id][subject].append(grade)

    # Вычисляем средние оценки
    data = {}
    for enrollment_id in enrollment_ids:
        data[enrollment_id] = {}
        for subj in subjects:
            grades_list = grades_by_student_subject[enrollment_id].get(subj, [])
            if grades_list:
                avg = sum(grades_list) / len(grades_list)
                data[enrollment_id][subj] = {
                    "average": round(avg, 2),
                    "grades": grades_list,
                    "count": len(grades_list),
                }
            else:
                data[enrollment_id][subj] = {"average": None, "grades": [], "count": 0}

    return {"students": students, "subjects": subjects, "data": data}


@router.get("/classes/{class_id}/journal/export/attendance")
def export_attendance_excel(
    class_id: str,
    user: dict = require_role("teacher", "admin"),
):
    """
    Экспорт журнала посещаемости в Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    sb = get_supabase()

    # Используем существующую функцию для получения данных
    journal_data = class_journal_by_dates(class_id, user)
    students = journal_data["students"]
    dates = journal_data["dates"]
    data = journal_data["data"]

    # Получаем название класса
    cls = sb.table("classes").select("name").eq("id", class_id).limit(1).execute().data
    class_name = cls[0].get("name") if cls else "Class"

    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    # Заголовки
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.cell(1, 1, "Ученик").fill = header_fill
    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, dt in enumerate(dates, start=2):
        cell = ws.cell(1, col_idx, dt)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные учеников
    for row_idx, student in enumerate(students, start=2):
        ws.cell(row_idx, 1, student["name"])

        student_data = data.get(student["id"], {})
        for col_idx, dt in enumerate(dates, start=2):
            day_data = student_data.get(dt, {})
            present = day_data.get("present")
            grade = day_data.get("grade")

            if grade:
                ws.cell(row_idx, col_idx, grade)
            elif present is True:
                ws.cell(row_idx, col_idx, "✓")
            elif present is False:
                ws.cell(row_idx, col_idx, "✗")

    # Авто-ширина колонок
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, len(dates) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{class_name}_poseschaiemost.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/classes/{class_id}/journal/export/grades")
def export_grades_excel(
    class_id: str,
    user: dict = require_role("teacher", "admin"),
):
    """
    Экспорт журнала оценок по предметам в Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    sb = get_supabase()

    # Используем существующую функцию для получения данных
    journal_data = class_journal_by_subject(class_id, user)
    students = journal_data["students"]
    subjects = journal_data["subjects"]
    data = journal_data["data"]

    # Получаем название класса
    cls = sb.table("classes").select("name").eq("id", class_id).limit(1).execute().data
    class_name = cls[0].get("name") if cls else "Class"

    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Оценки"

    # Заголовки
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws.cell(1, 1, "Ученик").fill = header_fill
    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, subject in enumerate(subjects, start=2):
        cell = ws.cell(1, col_idx, subject)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Добавляем колонку для средней оценки
    col_avg = len(subjects) + 2
    cell = ws.cell(1, col_avg, "Средний балл")
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные учеников
    for row_idx, student in enumerate(students, start=2):
        ws.cell(row_idx, 1, student["name"])

        student_data = data.get(student["id"], {})
        all_averages = []

        for col_idx, subject in enumerate(subjects, start=2):
            subj_data = student_data.get(subject, {})
            avg = subj_data.get("average")
            grades_list = subj_data.get("grades", [])

            if avg is not None:
                # Показываем среднюю оценку и все оценки в комментарии
                cell = ws.cell(row_idx, col_idx, avg)
                cell.comment = openpyxl.comments.Comment(
                    f"Оценки: {', '.join(map(str, grades_list))}", "System"
                )
                all_averages.append(avg)

        # Общий средний балл
        if all_averages:
            overall_avg = sum(all_averages) / len(all_averages)
            ws.cell(row_idx, col_avg, round(overall_avg, 2))

    # Авто-ширина колонок
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, len(subjects) + 3):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{class_name}_ocenki.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
