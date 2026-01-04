from __future__ import annotations

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import List

from app.core.deps import require_role
from app.db.supabase_client import get_supabase

router = APIRouter()


class SubjectTopicIn(BaseModel):
    topic_number: int
    topic_name: str
    lecture_hours: float = 0
    seminar_hours: float = 0
    practical_hours: float = 0
    exam_hours: float = 0
    description: str | None = None


class SubjectTopicOut(BaseModel):
    id: str
    subject_id: str
    topic_number: int
    topic_name: str
    lecture_hours: float
    seminar_hours: float
    practical_hours: float
    exam_hours: float
    total_hours: float
    description: str | None


@router.get("/subjects/{subject_id}/topics")
def list_subject_topics(subject_id: str, user: dict = require_role("admin", "manager", "teacher")):
    """Получить все темы учебного плана для предмета"""
    sb = get_supabase()
    
    # Verify subject exists
    subj = sb.table("subjects").select("id,name").eq("id", subject_id).limit(1).execute().data
    if not subj:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    topics = (
        sb.table("subject_topics")
        .select("*")
        .eq("subject_id", subject_id)
        .order("topic_number")
        .execute()
        .data or []
    )
    
    # Calculate totals
    total_lecture = sum(float(t.get("lecture_hours", 0)) for t in topics)
    total_seminar = sum(float(t.get("seminar_hours", 0)) for t in topics)
    total_practical = sum(float(t.get("practical_hours", 0)) for t in topics)
    total_exam = sum(float(t.get("exam_hours", 0)) for t in topics)
    total_all = total_lecture + total_seminar + total_practical + total_exam
    
    return {
        "subject": {"id": subj[0]["id"], "name": subj[0]["name"]},
        "topics": topics,
        "totals": {
            "lecture_hours": total_lecture,
            "seminar_hours": total_seminar,
            "practical_hours": total_practical,
            "exam_hours": total_exam,
            "total_hours": total_all
        }
    }


@router.post("/subjects/{subject_id}/topics")
def create_subject_topic(subject_id: str, payload: SubjectTopicIn, user: dict = require_role("admin", "manager")):
    """Создать тему в учебном плане"""
    sb = get_supabase()
    
    # Verify subject exists
    subj = sb.table("subjects").select("id").eq("id", subject_id).limit(1).execute().data
    if not subj:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    # Check if topic_number already exists
    existing = (
        sb.table("subject_topics")
        .select("id")
        .eq("subject_id", subject_id)
        .eq("topic_number", payload.topic_number)
        .limit(1)
        .execute()
        .data
    )
    if existing:
        raise HTTPException(status_code=409, detail="Topic number already exists for this subject")
    
    resp = sb.table("subject_topics").insert({
        "subject_id": subject_id,
        "topic_number": payload.topic_number,
        "topic_name": payload.topic_name,
        "lecture_hours": payload.lecture_hours,
        "seminar_hours": payload.seminar_hours,
        "practical_hours": payload.practical_hours,
        "exam_hours": payload.exam_hours,
        "description": payload.description
    }).execute()
    
    return {"topic": resp.data[0] if resp.data else resp.data}


@router.put("/subjects/{subject_id}/topics/{topic_id}")
def update_subject_topic(subject_id: str, topic_id: str, payload: SubjectTopicIn, user: dict = require_role("admin", "manager")):
    """Обновить тему в учебном плане"""
    sb = get_supabase()
    
    resp = sb.table("subject_topics").update({
        "topic_number": payload.topic_number,
        "topic_name": payload.topic_name,
        "lecture_hours": payload.lecture_hours,
        "seminar_hours": payload.seminar_hours,
        "practical_hours": payload.practical_hours,
        "exam_hours": payload.exam_hours,
        "description": payload.description,
        "updated_at": "now()"
    }).eq("id", topic_id).eq("subject_id", subject_id).execute()
    
    if not resp.data:
        raise HTTPException(status_code=404, detail="Topic not found")
    
    return {"topic": resp.data[0]}


@router.delete("/subjects/{subject_id}/topics/{topic_id}")
def delete_subject_topic(subject_id: str, topic_id: str, user: dict = require_role("admin", "manager")):
    """Удалить тему из учебного плана"""
    sb = get_supabase()
    
    sb.table("subject_topics").delete().eq("id", topic_id).eq("subject_id", subject_id).execute()
    
    return {"ok": True}


@router.post("/subjects/{subject_id}/topics/bulk-update")
def bulk_update_topics(subject_id: str, payload: List[SubjectTopicIn], user: dict = require_role("admin", "manager")):
    """Массовое обновление тем (замена всех тем)"""
    sb = get_supabase()
    
    # Verify subject exists
    subj = sb.table("subjects").select("id").eq("id", subject_id).limit(1).execute().data
    if not subj:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    # Delete all existing topics
    sb.table("subject_topics").delete().eq("subject_id", subject_id).execute()
    
    # Insert new topics
    if payload:
        topics_to_insert = [{
            "subject_id": subject_id,
            "topic_number": t.topic_number,
            "topic_name": t.topic_name,
            "lecture_hours": t.lecture_hours,
            "seminar_hours": t.seminar_hours,
            "practical_hours": t.practical_hours,
            "exam_hours": t.exam_hours,
            "description": t.description
        } for t in payload]
        
        sb.table("subject_topics").insert(topics_to_insert).execute()
    
    return {"ok": True, "count": len(payload)}


@router.get("/subjects/{subject_id}/topics/export")
def export_subject_topics_excel(subject_id: str, user: dict = require_role("admin", "manager", "teacher")):
    """Экспорт учебного плана в Excel"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse
    
    sb = get_supabase()
    
    # Get subject info
    subj = sb.table("subjects").select("id,name,code").eq("id", subject_id).limit(1).execute().data
    if not subj:
        raise HTTPException(status_code=404, detail="Subject not found")
    
    subject_name = subj[0].get("name", "Предмет")
    subject_code = subj[0].get("code", "")
    
    # Get topics
    topics = (
        sb.table("subject_topics")
        .select("*")
        .eq("subject_id", subject_id)
        .order("topic_number")
        .execute()
        .data or []
    )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Учебный план"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"УЧЕБНЫЙ ПЛАН: {subject_name}"
    if subject_code:
        title_cell.value += f" ({subject_code})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        "№",
        "Тема",
        "Лекции (ч)",
        "Семинары (ч)",
        "Практика (ч)",
        "Экзамен (ч)",
        "Всего (ч)",
        "Описание"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    total_lecture = 0
    total_seminar = 0
    total_practical = 0
    total_exam = 0
    total_all = 0
    
    for row_num, topic in enumerate(topics, 4):
        lecture = float(topic.get("lecture_hours", 0))
        seminar = float(topic.get("seminar_hours", 0))
        practical = float(topic.get("practical_hours", 0))
        exam = float(topic.get("exam_hours", 0))
        total = float(topic.get("total_hours", 0))
        
        total_lecture += lecture
        total_seminar += seminar
        total_practical += practical
        total_exam += exam
        total_all += total
        
        row_data = [
            topic.get("topic_number"),
            topic.get("topic_name"),
            lecture,
            seminar,
            practical,
            exam,
            total,
            topic.get("description", "")
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            if col_num > 2 and col_num < 8:  # Number columns
                cell.alignment = Alignment(horizontal="center")
    
    # Totals row
    totals_row = len(topics) + 4
    totals_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    totals_font = Font(bold=True)
    
    ws.cell(row=totals_row, column=1).value = ""
    ws.cell(row=totals_row, column=2).value = "ИТОГО:"
    ws.cell(row=totals_row, column=2).font = totals_font
    ws.cell(row=totals_row, column=3).value = total_lecture
    ws.cell(row=totals_row, column=4).value = total_seminar
    ws.cell(row=totals_row, column=5).value = total_practical
    ws.cell(row=totals_row, column=6).value = total_exam
    ws.cell(row=totals_row, column=7).value = total_all
    
    for col_num in range(1, 9):
        cell = ws.cell(row=totals_row, column=col_num)
        cell.fill = totals_fill
        cell.font = totals_font
        cell.border = thin_border
        if col_num > 2:
            cell.alignment = Alignment(horizontal="center")
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"uchebnyy_plan_{subject_code or subject_id}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
